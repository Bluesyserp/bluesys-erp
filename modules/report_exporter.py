# modules/report_exporter.py
import openpyxl
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import colors
from datetime import datetime

# --- 1. IMPORTAÇÕES FALTANTES ---
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
# --- FIM DA CORREÇÃO ---

# --- 1. EXPORTADOR XLSX (EXCEL) ---

def export_to_xlsx(headers, data, parent_widget):
    """
    Exporta os dados da tabela para um arquivo Excel (.xlsx).
    """
    try:
        # Pergunta ao usuário onde salvar
        default_name = f"Relatorio_Vendas_{datetime.now():%Y%m%d}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            parent_widget, 
            "Salvar Relatório Excel", 
            default_name, 
            "Excel Files (*.xlsx)"
        )
        
        if not save_path:
            return # Usuário cancelou

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Vendas"
        
        # Define o estilo do Cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Escreve os Cabeçalhos
        ws.append(headers)
        for cell in ws[1]: # Itera sobre a linha 1
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        # Escreve os Dados
        for row in data:
            ws.append(row)
            
        # Ajusta a largura das colunas
        for i, column_cells in enumerate(ws.columns):
            # +5 para dar um respiro
            length = max(len(str(cell.value) or "") for cell in column_cells) + 5
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = length

        wb.save(save_path)
        QMessageBox.information(parent_widget, "Sucesso", f"Relatório salvo com sucesso em:\n{save_path}")

    except Exception as e:
        QMessageBox.critical(parent_widget, "Erro ao Exportar XLSX", f"Falha ao gerar Excel: {e}")

# --- 2. EXPORTADOR PDF (A4) ---

def export_to_pdf(headers, data, title, parent_widget):
    """
    Exporta os dados da tabela para um PDF A4 "bonito" em modo paisagem.
    """
    try:
        default_name = f"Relatorio_Vendas_{datetime.now():%Y%m%d}.pdf"
        save_path, _ = QFileDialog.getSaveFileName(
            parent_widget, 
            "Salvar Relatório PDF", 
            default_name, 
            "PDF Files (*.pdf)"
        )
        
        if not save_path:
            return

        # Configura o documento A4 em modo Paisagem (landscape)
        doc = SimpleDocTemplate(save_path, pagesize=landscape(A4),
                                rightMargin=inch/2, leftMargin=inch/2,
                                topMargin=inch/2, bottomMargin=inch/2)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        style_title = styles['h1']
        style_title.alignment = TA_CENTER # <-- Agora funciona
        story.append(Paragraph(title, style_title))
        
        # Data
        style_normal = styles['Normal']
        style_normal.alignment = TA_CENTER # <-- Agora funciona
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", style_normal))

        # Adiciona um espaço
        story.append(Paragraph("<br/><br/>", style_normal))
        
        # Prepara dados para a tabela (Cabeçalho + dados)
        table_data = [headers] + data
        
        # Cria a Tabela
        t = Table(table_data, repeatRows=1) # Repete o cabeçalho em novas páginas
        
        # Estilo da Tabela
        style_table = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0078D7")), # Cor do cabeçalho
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('BOX', (0,0), (-1,-1), 0.25, colors.black),
            ('GRID', (0,0), (-1,-1), 0.25, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ])
        t.setStyle(style_table)
        
        story.append(t)
        
        # Constrói o PDF
        doc.build(story)
        QMessageBox.information(parent_widget, "Sucesso", f"Relatório salvo com sucesso em:\n{save_path}")

    except Exception as e:
        QMessageBox.critical(parent_widget, "Erro ao Exportar PDF", f"Falha ao gerar PDF: {e}")